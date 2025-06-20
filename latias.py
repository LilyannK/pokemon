import requests, pandas as pd, math
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import tempfile

API_KEY = "ccf311da-fbd7-48be-868d-a210163adc7a"
URL     = "https://api.pokemontcg.io/v2/cards"
PAGE_SIZE = 250  # max per page

def fetch_latias():
    headers = {"X-Api-Key": API_KEY}
    params  = {"q": 'name:"Latias"', "pageSize": PAGE_SIZE}
    first   = requests.get(URL, headers=headers, params=params).json()
    total   = first["totalCount"]
    pages   = math.ceil(total / PAGE_SIZE)
    cards   = first["data"]

    # grab remaining pages 
    for page in range(2, pages + 1):
        params["page"] = page
        cards.extend(requests.get(URL, headers=headers, params=params).json()["data"])
    return cards
def fetch_latios():
    headers = {"X-Api-Key": API_KEY}
    params  = {"q": 'name:"Latios"', "pageSize": PAGE_SIZE}
    first   = requests.get(URL, headers=headers, params=params).json()
    total   = first["totalCount"]
    pages   = math.ceil(total / PAGE_SIZE)
    cards   = first["data"]

    # grab remaining pages 
    for page in range(2, pages + 1):
        params["page"] = page
        cards.extend(requests.get(URL, headers=headers, params=params).json()["data"])
    return cards

def cards_to_df(cards):
    keep = ("id", "name", "set", "number", "rarity", "supertype", "subtypes",
            "hp", "types", "artist", "tcgplayer")
    rows = []

    for c in cards:
        rows.append({
            "id"        : c["id"],
            "set"       : c["set"]["name"],
            "set_code"  : c["set"]["id"],
            "number"    : c["number"],
            "rarity"    : c.get("rarity", ""),
            "variant"   : ", ".join(c.get("subtypes", [])),
           
            "img": c["images"]["large"]
        })
    return pd.DataFrame(rows).sort_values(["set_code", "number"])

cards = fetch_latias()
cards2 = fetch_latios()
df    = cards_to_df(cards) 
df2 =  cards_to_df(cards2)
#
df["Owned?"] = ""  # Leave blank to fill out manually
df2["Owned?"] = "" 
# Save initial Excel sheet
excel_file = "Latias_cards_with_images.xlsx"
df.to_excel(excel_file, index=False)

# Load workbook with openpyxl
wb = load_workbook(excel_file)
ws = wb.active

# Set column widths
ws.column_dimensions["A"].width = 18   # id
ws.column_dimensions["B"].width = 22   # set
ws.column_dimensions["K"].width = 12   # checkbox
ws.column_dimensions["L"].width = 18   # images

# Insert images in last column (L)
for i, url in enumerate(df["img"], start=2):  # Excel rows start at 2 (row 1 = header)
    try:
        img_data = requests.get(url).content
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(img_data)
            tmp.flush()
            img = XLImage(tmp.name)
            img.width = 100
            img.height = 140
            cell = f"L{i}"  # Column L for images
            ws.add_image(img, cell)
    except Exception as e:
        print(f"Failed to fetch image for row {i}: {e}")

# Save workbook with embedded images
wb.save(excel_file)
print("Latias card sheet created with images and checkboxes!")

# 5a. Local Excel workbook
df.to_excel("Latias_cards.xlsx", index=False)


excel_file2 = "Latios_cards_with_images.xlsx"
df2.to_excel(excel_file2, index=False)

# Load workbook with openpyxl
wb = load_workbook(excel_file2)
ws = wb.active

# Set column widths
ws.column_dimensions["A"].width = 18   # id
ws.column_dimensions["B"].width = 22   # set
ws.column_dimensions["K"].width = 12   # checkbox
ws.column_dimensions["L"].width = 18   # images

# Insert images in last column (L)
for i, url in enumerate(df2["img"], start=2):  # Excel rows start at 2 (row 1 = header)
    try:
        img_data = requests.get(url).content
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(img_data)
            tmp.flush()
            img = XLImage(tmp.name)
            img.width = 100
            img.height = 140
            cell = f"L{i}"  # Column L for images
            ws.add_image(img, cell)
    except Exception as e:
        print(f"Failed to fetch image for row {i}: {e}")

# Save workbook with embedded images
wb.save(excel_file2)
print("Lati0s card sheet created with images and checkboxes!")

# 5a. Local Excel workbook
df2.to_excel("Latios_cards.xlsx", index=False)
